import shutil
import tempfile

from fastapi.responses import FileResponse
import docx
import openpyxl as xl
from typing import List
from fastapi import APIRouter, Form, HTTPException, Path, UploadFile, File

router = APIRouter(prefix="/analyzer", tags=["Analyzer"])


def get_last_used_column(worksheet, row_number: int) -> int:
    """
    Find the last non-empty column in the specified row.
    Returns the column number (1-indexed).
    """
    max_column = worksheet.max_column

    # Start from the end and work backwards to find the last non-empty cell
    for col in range(max_column, 0, -1):
        if worksheet.cell(row=row_number, column=col).value is not None:
            return col

    return 0  # No data found


def process_word_documents(
    word_files: List[UploadFile],
    excel_file: UploadFile,
    sheet_name: str,
    words_column: str,
    words_start_row: int,
):
    """
    Process multiple Word documents and count word occurrences in Excel.

    Args:
        word_files: List of Word documents to process
        excel_file: Excel file containing the words to search for
        sheet_name: Name of the Excel sheet
        words_column: Column letter where words are listed (e.g., 'A')
        words_start_row: Starting row number for the words list
    """

    # Create temporary directory for file operations
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)

        # Save Excel file temporarily
        excel_path = temp_path / "workbook.xlsx"
        with open(excel_path, "wb") as buffer:
            shutil.copyfileobj(excel_file.file, buffer)

        # Load Excel workbook
        wb = xl.load_workbook(excel_path)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}",
            )

        ws = wb[sheet_name]

        # Find the last used column in the header row
        last_used_column = get_last_used_column(ws, words_start_row)

        # Start writing results from the next column
        result_start_column = last_used_column + 1

        # Convert column letter to number for word lookup
        words_column_num = ord(words_column.upper()) - ord("A") + 1

        # Process each Word document
        for doc_index, word_file in enumerate(word_files):
            # Save Word file temporarily
            word_path = temp_path / f"document_{doc_index}.docx"
            with open(word_path, "wb") as buffer:
                shutil.copyfileobj(word_file.file, buffer)

            # Read Word document content
            doc = docx.Document(word_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text.lower() + "."

            # Calculate the column for this document's results
            current_result_column = result_start_column + doc_index

            # Write document name in the header row
            ws.cell(row=words_start_row, column=current_result_column).value = Path(
                word_file.filename
            ).stem

            # Process each word in the Excel column
            current_row = words_start_row + 1
            while True:
                word_cell = ws.cell(row=current_row, column=words_column_num)
                word = word_cell.value

                if word is None:
                    break

                # Count occurrences (case-insensitive)
                word_lower = word.lower()
                count = text.count(word_lower)

                # Write count to Excel
                ws.cell(row=current_row, column=current_result_column).value = count

                current_row += 1

        # Save the modified workbook
        output_path = temp_path / "result.xlsx"
        wb.save(output_path)

        return output_path


@router.post("/process_unigramm")
async def process_documents(
    word_files: List[UploadFile] = File(..., description="Word documents to process"),
    excel_file: UploadFile = File(..., description="Excel file with words list"),
    sheet_name: str = Form(..., description="Excel sheet name"),
    words_column: str = Form(
        ..., description="Column letter where words are listed (e.g., 'A')"
    ),
    words_start_row: int = Form(..., description="Row number where words start"),
):
    """
    Process multiple Word documents and count word occurrences.

    Returns the modified Excel file with results.
    """

    # Validate file types
    for word_file in word_files:
        if not word_file.filename.endswith((".doc", ".docx")):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid Word file: {word_file.filename}. Only .doc and .docx files are allowed.",
            )

    if not excel_file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file: {excel_file.filename}. Only .xlsx files are allowed.",
        )

    try:
        # Process the files
        output_path = process_word_documents(
            word_files=word_files,
            excel_file=excel_file,
            sheet_name=sheet_name,
            words_column=words_column,
            words_start_row=words_start_row,
        )

        # Create a temporary copy of the result to return
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            shutil.copy(output_path, tmp.name)
            result_path = tmp.name

        return FileResponse(
            path=result_path,
            filename="processed_result.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")
